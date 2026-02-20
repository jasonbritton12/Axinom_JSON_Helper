try:
    import openpyxl
except ModuleNotFoundError:
    raise SystemExit(
        "Missing dependency: openpyxl\n"
        "Install with: python3 -m pip install openpyxl"
    )
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Protection

def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ingest Metadata"

    # --- Configuration ---
    
    # Dropdown Options
    ASSET_TYPES = '"MOVIE,EPISODE,SEASON,SHOW,CHANNEL"'
    VIDEO_PROFILES = '"nDRM (HLS),DRM (DASH & HLS)"'
    # Simplified country list for validation example (Excel has length limits for inline lists)
    # Users can still type others, but these will be suggested
    COUNTRIES = '"US,CA,GB,DE,FR,AU"' 

    # Headers structure: (Column Name, Description/Note, Width)
    headers = [
        ("Asset Type", "Required. Select from Dropdown.", 15),
        ("External ID", "Required. Unique Identifier.", 25),
        ("Title", "Required. Asset Title.", 30),
        ("Original Title", "Optional.", 25),
        ("Description", "Full description.", 40),
        ("Synopsis", "Short summary.", 40),
        ("Released Date", "YYYY-MM-DD", 15),
        ("Studio", "Studio Name", 20),
        ("Season/Ep Number", "Integer. For Seasons/Episodes.", 18),
        ("Parent External ID", "Link to Show/Season.", 25),
        ("Genres", "Comma separated (e.g. Action, Sci-Fi)", 30),
        ("Tags", "Comma separated", 25),
        ("Cast", "Comma separated", 30),
        ("Production Countries", "Comma separated ISO codes (US, CA)", 25),
        ("License Start (UTC)", "YYYY-MM-DD HH:MM", 20),
        ("License End (UTC)", "YYYY-MM-DD HH:MM", 20),
        ("License Countries", "Comma separated ISO codes", 25),
        ("Video Source", "Path (e.g. input/movie.mp4)", 30),
        ("Video Profile", "Select from Dropdown", 20),
        ("Cover Image", "Path to cover image", 30),
        ("Teaser Image", "Path to teaser image", 30),
    ]

    # --- Header Styling ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid") # Blue
    
    for col_idx, (header_name, note, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
        # Add comment/note as a tooltip substitute if needed, or just rely on column headers
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # --- Data Validation ---

    # 1. Asset Type (Column A)
    dv_type = DataValidation(type="list", formula1=ASSET_TYPES, allow_blank=False)
    dv_type.error = 'Your entry is not in the list'
    dv_type.errorTitle = 'Invalid Entry'
    ws.add_data_validation(dv_type)
    dv_type.add("A2:A1000")

    # 2. Production Countries (Column N) - Warning only, as it's comma sep, tough to validate strictly in Excel
    # We will just add a tip in the header or trust the user instructions.
    
    # 3. Video Profile (Column S)
    dv_profile = DataValidation(type="list", formula1=VIDEO_PROFILES, allow_blank=True)
    ws.add_data_validation(dv_profile)
    dv_profile.add("S2:S1000")

    # --- Freeze Panes ---
    ws.freeze_panes = "A2"

    # --- Save ---
    filename = "ingest_template.xlsx"
    wb.save(filename)
    print(f"Successfully generated '{filename}'.")

if __name__ == "__main__":
    create_template()
