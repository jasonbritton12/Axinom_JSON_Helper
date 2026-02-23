#!/usr/bin/env python3
"""Generate Axinom ingest Excel template v1.2.0 with validation and formatting."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.dimensions import ColumnDimension


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = ROOT / "docs" / "reference" / "axinom_ingest_template_v1_2_0.xlsx"
SHEET_NAME = "Ingest Metadata"
LIST_SHEET_NAME = "_lists"
MAX_ROWS = 1000

ASSET_TYPES = ["MOVIE", "TVSHOW", "SEASON", "EPISODE"]
VIDEO_PROFILES = ["DEFAULT", "nDRM (HLS)", "DRM (DASH & HLS)"]
LANGUAGE_TAGS = ["en-US", "en-GB", "es-ES", "fr-FR", "de-DE", "it-IT", "pt-BR", "sv-SE"]
COUNTRY_CODES = ["US", "CA", "GB", "DE", "FR", "AU", "ES", "IT", "SE", "NL"]

HEADERS = [
    ("Asset Type", 16, "Required. One of MOVIE, TVSHOW, SEASON, EPISODE."),
    ("External ID", 34, "Required for all asset types."),
    ("Title", 30, "Required for MOVIE, TVSHOW, and EPISODE."),
    ("Original Title", 26, "Optional."),
    ("Description", 42, "Optional long description."),
    ("Synopsis", 38, "Optional short synopsis."),
    ("Released Date", 16, "YYYY-MM-DD (Excel date also accepted)."),
    ("Studio", 22, "Optional."),
    ("Season/Ep Number", 18, "Required for SEASON and EPISODE. Integer > 0."),
    ("Parent External ID", 34, "Required for SEASON and EPISODE."),
    ("Genres", 28, "Comma-separated."),
    ("Tags", 24, "Comma-separated."),
    ("Cast", 34, "Comma-separated."),
    ("Production Countries", 26, "Comma-separated ISO2 country codes."),
    ("License Start (UTC)", 22, "UTC datetime. Excel date/time or text."),
    ("License End (UTC)", 22, "UTC datetime. Excel date/time or text."),
    ("License Countries", 26, "Comma-separated ISO2 country codes."),
    ("Video Source", 32, "Required when Video Profile is set."),
    ("Video Profile", 22, "Optional picklist."),
    ("Cover Image", 30, "Optional path."),
    ("Teaser Image", 30, "Optional path."),
    ("Language Tag", 18, "Required when any localized field is entered."),
    ("Localized Title", 28, "Optional localized title."),
    ("Localized Description", 34, "Optional localized description."),
    ("Localized Synopsis", 34, "Optional localized synopsis."),
    ("Trailer Source", 32, "Required when Trailer Profile is set."),
    ("Trailer Profile", 22, "Optional picklist."),
]

HEADER_INDEX = {name: idx + 1 for idx, (name, _, _) in enumerate(HEADERS)}

ALWAYS_REQUIRED_HEADERS = {
    "Asset Type",
    "External ID",
}

SOMETIMES_REQUIRED_HEADERS = {
    "Title",
    "Season/Ep Number",
    "Parent External ID",
    "Language Tag",
    "Video Source",
    "Trailer Source",
}

EXAMPLE_ROWS = [
    {
        "Asset Type": "MOVIE",
        "External ID": "M_example_movie",
        "Title": "Example Movie",
        "Description": "Example movie description.",
        "Synopsis": "Example movie synopsis.",
        "Released Date": "2025-01-01",
        "Studio": "Example Studio",
        "Genres": "Drama",
        "Production Countries": "US",
        "License Start (UTC)": "2025-01-01 12:00 AM",
        "License End (UTC)": "2026-01-01 11:59 PM",
        "License Countries": "US, CA",
        "Video Source": "input/ExampleMovie",
        "Video Profile": "DRM (DASH & HLS)",
        "Cover Image": "covers/example_movie.jpg",
        "Language Tag": "en-US",
        "Localized Title": "Example Movie",
        "Trailer Source": "trailers/example_movie_t1",
        "Trailer Profile": "DEFAULT",
    },
    {
        "Asset Type": "TVSHOW",
        "External ID": "S_example_show",
        "Title": "Example Show",
        "Description": "Example TV show description.",
        "Synopsis": "Example TV show synopsis.",
        "Released Date": "2024-01-01",
        "Studio": "Example Studio",
        "Genres": "Drama",
        "Production Countries": "US",
        "License Start (UTC)": "2024-01-01 12:00 AM",
        "License End (UTC)": "2099-12-31 11:59 PM",
        "License Countries": "US",
        "Video Profile": "DEFAULT",
        "Cover Image": "covers/example_show.jpg",
    },
    {
        "Asset Type": "SEASON",
        "External ID": "S_example_show_S1",
        "Title": "Example Show (Season 1)",
        "Season/Ep Number": "1",
        "Parent External ID": "S_example_show",
        "Description": "Season description.",
        "Released Date": "2024-01-01",
        "License Start (UTC)": "2024-01-01 12:00 AM",
        "License End (UTC)": "2099-12-31 11:59 PM",
        "License Countries": "US",
    },
    {
        "Asset Type": "EPISODE",
        "External ID": "S_example_show_S1_E1",
        "Title": "Episode 1",
        "Season/Ep Number": "1",
        "Parent External ID": "S_example_show_S1",
        "Description": "Episode description.",
        "Released Date": "2024-01-01",
        "License Start (UTC)": "2024-01-01 12:00 AM",
        "License End (UTC)": "2099-12-31 11:59 PM",
        "License Countries": "US",
        "Video Source": "input/ExampleShow/S1E1",
        "Video Profile": "nDRM (HLS)",
    },
]


def col_letter(col_index: int) -> str:
    value = col_index
    letters = ""
    while value > 0:
        value, remainder = divmod(value - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def add_list_validation(ws, formula: str, cell_range: str, *, allow_blank: bool = True, prompt: str = "") -> None:
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    dv.errorTitle = "Invalid value"
    dv.error = "Select a value from the drop-down list."
    if prompt:
        dv.promptTitle = "Allowed values"
        dv.prompt = prompt
        dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_decimal_int_validation(ws, cell_range: str) -> None:
    dv = DataValidation(type="whole", operator="greaterThan", formula1="0", allow_blank=True)
    dv.errorTitle = "Invalid number"
    dv.error = "Enter an integer greater than 0."
    dv.promptTitle = "Season/Episode Number"
    dv.prompt = "Required for SEASON and EPISODE. Must be an integer > 0."
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_date_validation(ws, cell_range: str, prompt: str) -> None:
    dv = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(1900,1,1)",
        formula2="DATE(2100,12,31)",
        allow_blank=True,
    )
    dv.errorTitle = "Invalid date"
    dv.error = "Enter a valid date (Excel date or YYYY-MM-DD text)."
    dv.promptTitle = "Date input"
    dv.prompt = prompt
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_datetime_validation(ws, cell_range: str, prompt: str) -> None:
    # Excel date/time serials pass. Text formats can still be pasted; helper validates on import.
    dv = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(1900,1,1)",
        formula2="DATE(2100,12,31)",
        allow_blank=True,
    )
    dv.errorTitle = "Invalid date/time"
    dv.error = "Enter an Excel date/time value or paste a helper-supported datetime string."
    dv.promptTitle = "UTC date/time"
    dv.prompt = prompt
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_country_hint_validation(ws, cell_range: str, list_sheet_name: str) -> None:
    formula = (
        f'=OR({cell_range.split(":")[0]}="",'
        f'COUNTIF({list_sheet_name}!$D$2:$D$11,{cell_range.split(":")[0]})>0,'
        f'ISNUMBER(SEARCH(",",{cell_range.split(":")[0]})))'
    )
    # Use a single-cell custom rule per column range by reusing top-left-relative formula.
    dv = DataValidation(type="custom", formula1=formula, allow_blank=True)
    dv.errorTitle = "Country codes"
    dv.error = "Use ISO2 codes. Single values should match the list; comma-separated entries are allowed."
    dv.promptTitle = "Country codes"
    dv.prompt = "Use ISO2 codes like US, CA. Comma-separated values are allowed."
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    thin = Side(style="thin", color="D0D7DE")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fills = {
        "always": PatternFill(fill_type="solid", fgColor="B42318"),
        "sometimes": PatternFill(fill_type="solid", fgColor="8250DF"),
        "optional": PatternFill(fill_type="solid", fgColor="1F6FEB"),
    }
    header_font = Font(bold=True, color="FFFFFF")

    for idx, (header, width, note) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border
        if header in ALWAYS_REQUIRED_HEADERS:
            cell.fill = fills["always"]
        elif header in SOMETIMES_REQUIRED_HEADERS:
            cell.fill = fills["sometimes"]
        else:
            cell.fill = fills["optional"]
        cell.comment = Comment(note, "Axinom Helper")
        ws.column_dimensions[col_letter(idx)] = ColumnDimension(ws, width=width)

    ws.row_dimensions[1].height = 34

    for row_idx, example in enumerate(EXAMPLE_ROWS, start=2):
        for header, value in example.items():
            col_idx = HEADER_INDEX[header]
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Basic wrapping for long text columns.
    for header in ("Description", "Synopsis", "Localized Description", "Localized Synopsis"):
        col_idx = HEADER_INDEX[header]
        for row in range(2, MAX_ROWS + 1):
            ws.cell(row=row, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")
    for row in range(2, len(EXAMPLE_ROWS) + 2):
        ws.row_dimensions[row].height = 42

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{col_letter(len(HEADERS))}{MAX_ROWS}"

    # Hidden list sheet for validation sources.
    lists = wb.create_sheet(LIST_SHEET_NAME)
    lists.sheet_state = "hidden"
    lists["A1"] = "Asset Types"
    for i, value in enumerate(ASSET_TYPES, start=2):
        lists[f"A{i}"] = value
    lists["B1"] = "Video Profiles"
    for i, value in enumerate(VIDEO_PROFILES, start=2):
        lists[f"B{i}"] = value
    lists["C1"] = "Language Tags"
    for i, value in enumerate(LANGUAGE_TAGS, start=2):
        lists[f"C{i}"] = value
    lists["D1"] = "Country Codes"
    for i, value in enumerate(COUNTRY_CODES, start=2):
        lists[f"D{i}"] = value

    # Data validation.
    add_list_validation(
        ws,
        f"={LIST_SHEET_NAME}!$A$2:$A${1 + len(ASSET_TYPES)}",
        f"A2:A{MAX_ROWS}",
        allow_blank=False,
        prompt="Choose a supported asset type.",
    )
    add_decimal_int_validation(ws, f"I2:I{MAX_ROWS}")
    add_date_validation(ws, f"G2:G{MAX_ROWS}", "Use Excel date or YYYY-MM-DD.")
    add_datetime_validation(ws, f"O2:O{MAX_ROWS}", "UTC start datetime; Excel date/time works best.")
    add_datetime_validation(ws, f"P2:P{MAX_ROWS}", "UTC end datetime; Excel date/time works best.")
    add_list_validation(
        ws,
        f"={LIST_SHEET_NAME}!$B$2:$B${1 + len(VIDEO_PROFILES)}",
        f"S2:S{MAX_ROWS}",
        prompt="Optional. If set, Video Source should also be set.",
    )
    add_list_validation(
        ws,
        f"={LIST_SHEET_NAME}!$B$2:$B${1 + len(VIDEO_PROFILES)}",
        f"AB2:AB{MAX_ROWS}",
        prompt="Optional. If set, Trailer Source should also be set.",
    )
    add_list_validation(
        ws,
        f"={LIST_SHEET_NAME}!$C$2:$C${1 + len(LANGUAGE_TAGS)}",
        f"V2:V{MAX_ROWS}",
        prompt="Required if any localized fields are populated.",
    )
    add_country_hint_validation(ws, f"N2:N{MAX_ROWS}", LIST_SHEET_NAME)
    add_country_hint_validation(ws, f"Q2:Q{MAX_ROWS}", LIST_SHEET_NAME)

    # Conditional formatting for required-but-blank fields.
    missing_fill = PatternFill(fill_type="solid", fgColor="FDECEC")
    missing_font = Font(color="B42318", bold=True)

    def add_required_blank_cf(col: str, formula: str) -> None:
        ws.conditional_formatting.add(
            f"{col}2:{col}{MAX_ROWS}",
            FormulaRule(formula=[formula], fill=missing_fill, font=missing_font),
        )

    add_required_blank_cf("A", '=LEN(TRIM($A2))=0')
    add_required_blank_cf("B", '=AND(LEN(TRIM($A2))>0,LEN(TRIM($B2))=0)')
    add_required_blank_cf("C", '=AND(OR($A2="MOVIE",$A2="TVSHOW",$A2="EPISODE"),LEN(TRIM($C2))=0)')
    add_required_blank_cf("I", '=AND(OR($A2="SEASON",$A2="EPISODE"),LEN(TRIM($I2))=0)')
    add_required_blank_cf("J", '=AND(OR($A2="SEASON",$A2="EPISODE"),LEN(TRIM($J2))=0)')
    add_required_blank_cf("R", '=AND(LEN(TRIM($S2))>0,LEN(TRIM($R2))=0)')
    add_required_blank_cf("V", '=AND(OR(LEN(TRIM($W2))>0,LEN(TRIM($X2))>0,LEN(TRIM($Y2))>0),LEN(TRIM($V2))=0)')
    add_required_blank_cf("AA", '=AND(LEN(TRIM($AB2))>0,LEN(TRIM($AA2))=0)')

    # Light body fill for the example rows only.
    for row in range(2, len(EXAMPLE_ROWS) + 2):
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=row, column=col).fill = PatternFill(fill_type="solid", fgColor="F6F8FA")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
