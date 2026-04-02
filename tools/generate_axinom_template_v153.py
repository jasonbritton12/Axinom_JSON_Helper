#!/usr/bin/env python3
"""Generate Axinom ingest Excel template v1.5.3 with expanded profiles."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = ROOT / "docs" / "reference" / "axinom_ingest_template_v1_5_3.xlsx"
SHEET_NAME = "Ingest Metadata"
LIST_SHEET_NAME = "_lists"
MAX_ROWS = 1000

ASSET_TYPES = ["MOVIE", "TVSHOW", "SEASON", "EPISODE", "TRAILER", "EXTRA"]
VIDEO_PROFILES = [
    "CMAF_File_Non-DRM",
    "CMAF_File_DRM",
    "CMAF_File_Non-DRM_SD",
    "CMAF_File_DRM_SD",
]
COUNTRY_CODES = ["US", "CA", "GB", "DE", "FR", "AU", "ES", "IT", "SE", "NL"]

HEADERS = [
    ("Asset Type", 15.0, "Required. One of MOVIE, TVSHOW, SEASON, EPISODE, TRAILER, EXTRA."),
    ("External ID", 78.0, "Required for all asset types."),
    ("Title", 54.33203125, "Required for MOVIE, TVSHOW, EPISODE, TRAILER, and EXTRA."),
    ("Original Title", 40.1640625, "Optional."),
    ("Description", 40.0, "Optional long description."),
    ("Synopsis", 13.0, "Optional short synopsis."),
    ("Released Date", 15.0, "YYYY-MM-DD (Excel date also accepted)."),
    ("Studio", 22.1640625, "Optional."),
    ("Season/Ep Number", 18.0, "Required for SEASON and EPISODE. Integer > 0."),
    ("Parent Type", 18.0, "Helper field. Required for TRAILER and EXTRA. Use TVSHOW for SEASON parents and SEASON for EPISODE parents if you want explicit validation."),
    ("Parent External ID", 74.1640625, "Required for SEASON, EPISODE, TRAILER, and EXTRA."),
    ("Genres", 30.0, "Comma-separated."),
    ("Tags", 25.0, "Comma-separated."),
    ("Cast", 30.0, "Comma-separated."),
    ("Production Countries", 25.0, "Comma-separated ISO2 country codes."),
    ("License Start (UTC)", 24.33203125, "UTC datetime. Excel date/time or text."),
    ("License End (UTC)", 20.0, "UTC datetime. Excel date/time or text."),
    ("License Countries", 25.0, "Comma-separated ISO2 country codes."),
    ("Video Source", 62.1640625, "Required for MOVIE, EPISODE, TRAILER, and EXTRA in minimum ingest flow."),
    ("Video Profile", 20.0, "Required for MOVIE, EPISODE, TRAILER, and EXTRA in minimum ingest flow."),
    ("Trailer Source", 62.1640625, "Optional nested trailer source. Required when Trailer Profile is set."),
    ("Trailer Profile", 20.0, "Optional picklist."),
    ("Cover Image", 30.0, "Optional path."),
    ("Teaser Image", 13.0, "Optional path."),
]

HEADER_INDEX = {name: idx + 1 for idx, (name, _, _) in enumerate(HEADERS)}

ALWAYS_REQUIRED_HEADERS = {
    "Asset Type",
    "External ID",
}

SOMETIMES_REQUIRED_HEADERS = {
    "Title",
    "Season/Ep Number",
    "Parent Type",
    "Parent External ID",
    "Video Source",
    "Video Profile",
    "Trailer Source",
}

EXAMPLE_ROWS = [
    {
        "Asset Type": "MOVIE",
        "External ID": "M_example_movie",
        "Title": "Example Movie",
        "Original Title": "Example Movie",
        "Description": "Example movie description.",
        "Synopsis": "Example movie synopsis.",
        "Released Date": "2025-01-01",
        "Studio": "Example Studio",
        "Genres": "Drama",
        "Tags": "featured",
        "Cast": "Actor A, Actor B",
        "Production Countries": "US",
        "License Start (UTC)": "2025-01-01 09:00 PM",
        "License End (UTC)": "2026-01-01 11:59 PM",
        "License Countries": "US, CA",
        "Video Source": "input/ExampleMovie",
        "Video Profile": "CMAF_File_DRM",
        "Trailer Source": "trailers/example_movie_t1",
        "Trailer Profile": "CMAF_File_Non-DRM",
        "Cover Image": "covers/example_movie.jpg",
        "Teaser Image": "teasers/example_movie.jpg",
    },
    {
        "Asset Type": "TVSHOW",
        "External ID": "S_example_show",
        "Title": "Example Show",
        "Original Title": "Example Show",
        "Description": "Example TV show description.",
        "Synopsis": "Example TV show synopsis.",
        "Released Date": "2024-01-01",
        "Studio": "Example Studio",
        "Genres": "Drama",
        "Tags": "series",
        "Cast": "Actor C, Actor D",
        "Production Countries": "US",
        "License Start (UTC)": "2024-01-01 09:00 PM",
        "License End (UTC)": "2099-12-31 11:59 PM",
        "License Countries": "US",
        "Trailer Source": "trailers/example_show_t1",
        "Trailer Profile": "CMAF_File_Non-DRM",
        "Cover Image": "covers/example_show.jpg",
        "Teaser Image": "teasers/example_show.jpg",
    },
    {
        "Asset Type": "SEASON",
        "External ID": "S_example_show_S1",
        "Title": "Example Show (Season 1)",
        "Season/Ep Number": "1",
        "Parent Type": "TVSHOW",
        "Parent External ID": "S_example_show",
        "Description": "Season description.",
        "Released Date": "2024-01-01",
        "License Start (UTC)": "2024-01-01 09:00 PM",
        "License End (UTC)": "2099-12-31 11:59 PM",
        "License Countries": "US",
        "Cover Image": "covers/example_show_s1.jpg",
    },
    {
        "Asset Type": "EPISODE",
        "External ID": "S_example_show_S1_E1",
        "Title": "Episode 1",
        "Season/Ep Number": "1",
        "Parent Type": "SEASON",
        "Parent External ID": "S_example_show_S1",
        "Description": "Episode description.",
        "Released Date": "2024-01-01",
        "License Start (UTC)": "2024-01-01 09:00 PM",
        "License End (UTC)": "2099-12-31 11:59 PM",
        "License Countries": "US",
        "Video Source": "input/ExampleShow/S1E1",
        "Video Profile": "CMAF_File_Non-DRM",
        "Cover Image": "covers/example_show_s1e1.jpg",
    },
    {
        "Asset Type": "TRAILER",
        "External ID": "T_example_movie_t1",
        "Title": "Example Movie Trailer",
        "Parent Type": "MOVIE",
        "Parent External ID": "M_example_movie",
        "Description": "Trailer description.",
        "Released Date": "2025-01-01",
        "License Start (UTC)": "2025-01-01 09:00 PM",
        "License End (UTC)": "2026-01-01 11:59 PM",
        "License Countries": "US",
        "Video Source": "trailers/ExampleMovie/t1",
        "Video Profile": "CMAF_File_Non-DRM",
        "Cover Image": "covers/example_movie_trailer.jpg",
    },
    {
        "Asset Type": "EXTRA",
        "External ID": "X_example_show_bts",
        "Title": "Behind The Scenes",
        "Parent Type": "TVSHOW",
        "Parent External ID": "S_example_show",
        "Description": "Companion extra content.",
        "Released Date": "2024-01-15",
        "License Start (UTC)": "2024-01-15 09:00 PM",
        "License End (UTC)": "2099-12-31 11:59 PM",
        "License Countries": "US",
        "Video Source": "extras/ExampleShow/bts",
        "Video Profile": "CMAF_File_DRM",
        "Trailer Source": "trailers/example_show_bts_t1",
        "Trailer Profile": "CMAF_File_Non-DRM",
        "Cover Image": "covers/example_show_bts.jpg",
    },
]


def col_letter(col_index: int) -> str:
    value = col_index
    letters = ""
    while value > 0:
        value, remainder = divmod(value - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def add_list_validation(ws, formula: str, cell_range: str, *, allow_blank: bool = True, prompt: str = "") -> None:
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    dv.errorTitle = "Invalid value"
    dv.error = "Select a value from the drop-down list."
    if prompt:
        dv.promptTitle = "Allowed values"
        dv.prompt = prompt
        dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_integer_validation(ws, cell_range: str) -> None:
    dv = DataValidation(type="whole", operator="greaterThan", formula1="0", allow_blank=True)
    dv.errorTitle = "Invalid number"
    dv.error = "Enter an integer greater than 0."
    dv.promptTitle = "Season/Episode Number"
    dv.prompt = "Required for SEASON and EPISODE. Must be an integer greater than 0."
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_date_validation(ws, cell_range: str, prompt: str) -> None:
    dv = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(1900,1,1)",
        formula2="DATE(2100,12,31)",
        allow_blank=True,
    )
    dv.errorTitle = "Invalid date"
    dv.error = "Enter a valid date (Excel date or YYYY-MM-DD text)."
    dv.promptTitle = "Date input"
    dv.prompt = prompt
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_datetime_validation(ws, cell_range: str, prompt: str) -> None:
    dv = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(1900,1,1)",
        formula2="DATE(2100,12,31)",
        allow_blank=True,
    )
    dv.errorTitle = "Invalid date/time"
    dv.error = "Enter an Excel date/time value or paste a helper-supported datetime string."
    dv.promptTitle = "UTC date/time"
    dv.prompt = prompt
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_country_hint_validation(ws, cell_range: str, list_sheet_name: str) -> None:
    first_cell = cell_range.split(":")[0]
    formula = (
        f'=OR({first_cell}="",'
        f'COUNTIF({list_sheet_name}!$C$2:$C${1 + len(COUNTRY_CODES)},{first_cell})>0,'
        f'ISNUMBER(SEARCH(",",{first_cell})))'
    )
    dv = DataValidation(type="custom", formula1=formula, allow_blank=True)
    dv.errorTitle = "Country codes"
    dv.error = "Use ISO2 codes. Single values should match the list; comma-separated entries are allowed."
    dv.promptTitle = "Country codes"
    dv.prompt = "Use ISO2 codes like US, CA. Comma-separated values are allowed."
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_required_blank_cf(ws, header: str, formula: str) -> None:
    col = col_letter(HEADER_INDEX[header])
    missing_fill = PatternFill(fill_type="solid", fgColor="FDECEC")
    missing_font = Font(name="Poppins Regular", size=12, color="B42318", bold=True)
    ws.conditional_formatting.add(
        f"{col}2:{col}{MAX_ROWS}",
        FormulaRule(formula=[formula], fill=missing_fill, font=missing_font),
    )


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    header_fills = {
        "always": PatternFill(fill_type="solid", fgColor="B42318"),
        "sometimes": PatternFill(fill_type="solid", fgColor="8250DF"),
        "optional": PatternFill(fill_type="solid", fgColor="2563EB"),
    }
    header_font = Font(name="Poppins Regular", size=12, bold=True, color="FFFFFF")
    body_font = Font(name="Poppins Regular", size=12, bold=False, color="000000")

    for idx, (header, width, note) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if header in ALWAYS_REQUIRED_HEADERS:
            cell.fill = header_fills["always"]
        elif header in SOMETIMES_REQUIRED_HEADERS:
            cell.fill = header_fills["sometimes"]
        else:
            cell.fill = header_fills["optional"]
        cell.comment = Comment(note, "Axinom Helper")
        ws.column_dimensions[col_letter(idx)].width = width

    for row_idx, example in enumerate(EXAMPLE_ROWS, start=2):
        for header, value in example.items():
            col_idx = HEADER_INDEX[header]
            ws.cell(row=row_idx, column=col_idx, value=value)

    for row in range(2, MAX_ROWS + 1):
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=row, column=col).font = body_font

    for header in ("Description", "Synopsis"):
        col_idx = HEADER_INDEX[header]
        for row in range(2, MAX_ROWS + 1):
            ws.cell(row=row, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    for row in range(2, len(EXAMPLE_ROWS) + 2):
        ws.row_dimensions[row].height = 40

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{col_letter(len(HEADERS))}{MAX_ROWS}"

    lists = wb.create_sheet(LIST_SHEET_NAME)
    lists.sheet_state = "hidden"

    lists["A1"] = "Asset Types"
    for i, value in enumerate(ASSET_TYPES, start=2):
        lists[f"A{i}"] = value

    lists["B1"] = "Video Profiles"
    for i, value in enumerate(VIDEO_PROFILES, start=2):
        lists[f"B{i}"] = value

    lists["C1"] = "Country Codes"
    for i, value in enumerate(COUNTRY_CODES, start=2):
        lists[f"C{i}"] = value

    add_list_validation(
        ws,
        f"={LIST_SHEET_NAME}!$A$2:$A${1 + len(ASSET_TYPES)}",
        f"A2:A{MAX_ROWS}",
        allow_blank=False,
        prompt="Choose a supported asset type.",
    )
    add_integer_validation(ws, f"I2:I{MAX_ROWS}")
    add_list_validation(
        ws,
        f"={LIST_SHEET_NAME}!$A$2:$A${1 + len(ASSET_TYPES)}",
        f"J2:J{MAX_ROWS}",
        prompt="Parent Type helper. Required for TRAILER and EXTRA.",
    )
    add_date_validation(ws, f"G2:G{MAX_ROWS}", "Use Excel date or YYYY-MM-DD.")
    add_datetime_validation(ws, f"P2:P{MAX_ROWS}", "UTC start datetime; Excel date/time works best.")
    add_datetime_validation(ws, f"Q2:Q{MAX_ROWS}", "UTC end datetime; Excel date/time works best.")
    add_list_validation(
        ws,
        f"={LIST_SHEET_NAME}!$B$2:$B${1 + len(VIDEO_PROFILES)}",
        f"T2:T{MAX_ROWS}",
        prompt="Pick one of the supported video profiles.",
    )
    add_list_validation(
        ws,
        f"={LIST_SHEET_NAME}!$B$2:$B${1 + len(VIDEO_PROFILES)}",
        f"V2:V{MAX_ROWS}",
        prompt="Optional trailer profile.",
    )
    add_country_hint_validation(ws, f"O2:O{MAX_ROWS}", LIST_SHEET_NAME)
    add_country_hint_validation(ws, f"R2:R{MAX_ROWS}", LIST_SHEET_NAME)

    add_required_blank_cf(ws, "Asset Type", '=LEN(TRIM($A2))=0')
    add_required_blank_cf(ws, "External ID", '=AND(LEN(TRIM($A2))>0,LEN(TRIM($B2))=0)')
    add_required_blank_cf(
        ws,
        "Title",
        '=AND(OR($A2="MOVIE",$A2="TVSHOW",$A2="EPISODE",$A2="TRAILER",$A2="EXTRA"),LEN(TRIM($C2))=0)',
    )
    add_required_blank_cf(ws, "Season/Ep Number", '=AND(OR($A2="SEASON",$A2="EPISODE"),LEN(TRIM($I2))=0)')
    add_required_blank_cf(
        ws,
        "Parent Type",
        '=AND(OR($A2="TRAILER",$A2="EXTRA"),LEN(TRIM($J2))=0)',
    )
    add_required_blank_cf(
        ws,
        "Parent External ID",
        '=AND(OR($A2="SEASON",$A2="EPISODE",$A2="TRAILER",$A2="EXTRA"),LEN(TRIM($K2))=0)',
    )
    add_required_blank_cf(
        ws,
        "Video Source",
        '=AND(OR($A2="MOVIE",$A2="EPISODE",$A2="TRAILER",$A2="EXTRA",LEN(TRIM($T2))>0),LEN(TRIM($S2))=0)',
    )
    add_required_blank_cf(
        ws,
        "Video Profile",
        '=AND(OR($A2="MOVIE",$A2="EPISODE",$A2="TRAILER",$A2="EXTRA",LEN(TRIM($S2))>0),LEN(TRIM($T2))=0)',
    )
    add_required_blank_cf(ws, "Trailer Source", '=AND(LEN(TRIM($V2))>0,LEN(TRIM($U2))=0)')

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
